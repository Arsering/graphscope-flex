#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 18 16:44:14 2023
@author: yz
"""

import os
import sys
import re
import json
import numpy as np
from matplotlib import pyplot as plt
import matplotlib as mpl
import pandas as pd
from openpyxl import load_workbook
import multiprocessing
import math

CUR_FILE_PATH = os.path.dirname(__file__)
sys.path.append(CUR_FILE_PATH + '/../')
# from plt_mine import set_mpl

def set_mpl():
    mpl.rcParams.update(
        {
            'text.usetex': False,
            # 'font.sans-serif': 'Times New Roman',
            'mathtext.fontset': 'stix',
            'font.size': 20,
            'figure.figsize': (10.0, 10 * 0.618),
            'savefig.dpi': 10000,
            'axes.labelsize': 25,
            'axes.linewidth': 1.2,
            'xtick.labelsize': 20,
            'ytick.labelsize': 20,
            'legend.loc': 'upper right',
            'lines.linewidth': 3,
            'lines.markersize': 5,
            # 'axes.labelweight': 'bold'
        }
    )
        
def plt_figure_2(file_path):
    set_mpl()
    fig = plt.figure()
    ax1 = fig.subplots()
    plt.grid(True,  axis='both', linestyle='--', color='gray', linewidth=0.5)
    plt.xlabel('Worker Number')

    wb = load_workbook(file_path)
    # for row in wb['batching_IO_2G'].iter_rows(values_only=True):
    exp_types = ['batching_IO', 'sequential_IO']
    line_styles = ['-', '--']
    ax2 = ax1.twinx()
    plt.grid(True,  axis='both', linestyle='--', color='gray', linewidth=0.5)
    memory_size = 2
    for exp_type, line_style in zip(exp_types, line_styles):
        xs = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][1]][1:]
        ys_1 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][4]][1:]
        ys_2 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][2]][1:]

        ax1.plot(xs, ys_1, linestyle=line_style, color=plt.get_cmap('tab10')(0), label=exp_type+'+'+ str(memory_size*10) +'%')
        ax2.plot(xs, ys_2, linestyle=line_style, color=plt.get_cmap('tab10')(2), label=exp_type+'+'+ str(memory_size*10) +'%')

    ax1.set_ylabel('Latency (us)', color=plt.get_cmap('tab10')(0))
    ax1.set_ylim(0, 1300)
    ax1.set_yticks(range(0, 1250, 200), range(0, 1250, 200))
    ax1.tick_params(axis='y', colors=plt.get_cmap('tab10')(0))
    
    ax2.set_ylabel('Worker Throughput (GB)', color=plt.get_cmap('tab10')(2))
    ax2.set_ylim(0, 3.25)
    ax2.set_yticks(np.array(range(0, 31, 5))/10, np.array(range(0, 31, 5))/10)
    ax2.tick_params(axis='y', colors=plt.get_cmap('tab10')(2))

    plt.xlabel('Worker Number')
    plt.legend(loc='upper left')
    plt.savefig('figs/fig13-2g.pdf', bbox_inches='tight')


def plt_figure_5(file_path):
    set_mpl()
    fig = plt.figure()
    ax1 = fig.subplots()
    plt.grid(True,  axis='both', linestyle='--', color='gray', linewidth=0.5)
    plt.xlabel('Worker Number')

    wb = load_workbook(file_path)
    # for row in wb['batching_IO_2G'].iter_rows(values_only=True):
    exp_types = ['batching_IO', 'sequential_IO']
    line_styles = ['-', '--']
    ax2 = ax1.twinx()
    memory_size = 5
    for exp_type, line_style in zip(exp_types, line_styles):
        xs = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][1]][1:]
        ys_1 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][4]][1:]
        ys_2 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][2]][1:]

        ax1.plot(xs, ys_1, linestyle=line_style, color=plt.get_cmap('tab10')(0), label=exp_type+'+'+ str(memory_size*10) +'%')
        ax2.plot(xs, ys_2, linestyle=line_style, color=plt.get_cmap('tab10')(2), label=exp_type+'+'+ str(memory_size*10) +'%')

    ax1.set_ylabel('Latency (us)', color=plt.get_cmap('tab10')(0))
    ax1.set_ylim(0, 975)
    ax1.set_yticks(range(0, 910, 150), range(0, 910, 150))
    ax1.tick_params(axis='y', colors=plt.get_cmap('tab10')(0))
    
    ax2.set_ylabel('Worker Throughput (GB)', color=plt.get_cmap('tab10')(2))
    ax2.set_ylim(0, 16.25)
    ax2.set_yticks(np.array(range(0, 160, 25))/10, np.array(range(0, 160, 25))/10)
    ax2.tick_params(axis='y', colors=plt.get_cmap('tab10')(2))

    plt.xlabel('Worker Number')
    plt.legend(loc='upper left')
    plt.savefig('figs/fig13-5g.pdf', bbox_inches='tight')
    
    
def plt_figure_8(file_path):
    set_mpl()
    fig = plt.figure()
    ax1 = fig.subplots()
    plt.grid(True,  axis='both', linestyle='--', color='gray', linewidth=0.5)
    plt.xlabel('Worker Number')

    wb = load_workbook(file_path)
    # for row in wb['batching_IO_2G'].iter_rows(values_only=True):
    exp_types = ['batching_IO', 'sequential_IO']
    line_styles = ['-', '--']
    ax2 = ax1.twinx()
    memory_size = 8
    for exp_type, line_style in zip(exp_types, line_styles):
        xs = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][1]][1:]
        ys_1 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][4]][1:]
        ys_2 = [cell.value for cell in wb[exp_type+'_'+ str(memory_size) +'G'][2]][1:]

        ax1.plot(xs, ys_1, linestyle=line_style, color=plt.get_cmap('tab10')(0), label=exp_type+'+'+ str(memory_size*10) +'%')
        ax2.plot(xs, ys_2, linestyle=line_style, color=plt.get_cmap('tab10')(2), label=exp_type+'+'+ str(memory_size*10) +'%')

    ax1.set_ylabel('Latency (us)', color=plt.get_cmap('tab10')(0))
    ax1.set_ylim(0, 325)
    ax1.set_yticks(range(0, 350, 50), range(0, 350, 50))
    ax1.tick_params(axis='y', colors=plt.get_cmap('tab10')(0))
    
    ax2.set_ylabel('Worker Throughput (GB)', color=plt.get_cmap('tab10')(2))
    ax2.set_ylim(0, 16.25)
    ax2.set_yticks(np.array(range(0, 160, 25))/10, np.array(range(0, 160, 25))/10)
    ax2.tick_params(axis='y', colors=plt.get_cmap('tab10')(2))
    
    plt.legend(loc='upper left')
    plt.savefig('figs/fig13-8g.pdf', bbox_inches='tight')
    
if __name__ == "__main__":
    file_path = '/data/zhengyang/data/graphscope-flex/experiment_space/LDBC_SNB/python/data/figure_13/batching_IO-vs-sequential_IO.xlsx'
    
    plt_figure_8(file_path)
    plt_figure_5(file_path)
    plt_figure_2(file_path)

    print('\nwork finished')


